from unittest import case

from flask import (
    Flask,
    render_template,
    redirect,
    url_for,
    request,
    flash,
    send_file
)

from io import BytesIO

from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    login_required,
    current_user
)

from werkzeug.security import (
    check_password_hash,
    generate_password_hash
)

from config import Config
from models import *
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
# ==========================
# APP
# ==========================

app = Flask(__name__)
app.config.from_object(Config)

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


with app.app_context():
    db.create_all()
    create_admin()


# ==========================
# HOME
# ==========================

@app.route("/")
def home():

    if current_user.is_authenticated:
        return redirect(
            url_for("dashboard")
        )

    return redirect(
        url_for("login")
    )


# ==========================
# LOGIN
# ==========================

@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        user = User.query.filter_by(
            username=username
        ).first()

        if user and check_password_hash(
            user.password,
            password
        ):

            login_user(user)

            return redirect(
                url_for("dashboard")
            )

        flash(
            "Username atau password salah",
            "danger"
        )

    return render_template(
        "login.html"
    )


# ==========================
# LOGOUT
# ==========================

@app.route("/logout")
@login_required
def logout():

    logout_user()

    return redirect(
        url_for("login")
    )


# ==========================
# DASHBOARD
# ==========================

@app.route("/dashboard")
@login_required
def dashboard():

    # ADMIN
    if current_user.role == "admin":

        total_kegiatan = Kegiatan.query.count()
        total_ppl = User.query.filter_by(role="ppl").count()
        total_pml = User.query.filter_by(role="pml").count()
        total_bs = BlokSensus.query.count()

        kegiatan_progress = []

        tahun_filter = request.args.get(
    "tahun",
    type=int
)

        query = Kegiatan.query

        if tahun_filter:
            query = query.filter(
                Kegiatan.tahun == tahun_filter
            )

        kegiatan_list = query.order_by(
            db.case(
                (
                    Kegiatan.status == "Berjalan",
                    0
                ),
                else_=1
            ),
            Kegiatan.tanggal_mulai
        ).all()

        tahun_list = db.session.query(
            Kegiatan.tahun
        ).distinct().order_by(
            Kegiatan.tahun.desc()
        ).all()

        tahun_list = [x[0] for x in tahun_list]

        for kegiatan in kegiatan_list:

            bs_list = BlokSensus.query.filter_by(
                kegiatan_id=kegiatan.id
            ).all()

            total_target = sum(
                bs.target for bs in bs_list
            )

            total_realisasi = sum(
                bs.realisasi for bs in bs_list
            )

            total_diperiksa = sum(
                getattr(bs, "diperiksa", 0)
                for bs in bs_list
            )

            progress_ppl = 0
            progress_pml = 0

            if total_target > 0:

                progress_ppl = round(
                    total_realisasi * 100 / total_target,
                    1
                )

                progress_pml = round(
                    total_diperiksa * 100 / total_target,
                    1
                )

            kegiatan_progress.append({
                "id": kegiatan.id,
                "nama": kegiatan.nama,
                "target": total_target,
                "realisasi": total_realisasi,
                "diperiksa": total_diperiksa,
                "progress_ppl": progress_ppl,
                "progress_pml": progress_pml,
                "status": kegiatan.status
            })
        # CARD DASHBOARD

        kegiatan_berjalan = len([
            k for k in kegiatan_list
            if k.status == "Berjalan"
        ])

        total_target = 0
        total_realisasi = 0
        total_diperiksa = 0

        for kegiatan in kegiatan_list:

            bs_list = BlokSensus.query.filter_by(
                kegiatan_id=kegiatan.id
            ).all()

            total_target += sum(
                bs.target for bs in bs_list
            )

            total_realisasi += sum(
                bs.realisasi for bs in bs_list
            )

            total_diperiksa += sum(
                getattr(bs, "diperiksa", 0)
                for bs in bs_list
            )
  
        return render_template(
            "dashboard_admin.html",

            kegiatan_berjalan=kegiatan_berjalan,

            total_target=total_target,

            total_realisasi=total_realisasi,

            total_diperiksa=total_diperiksa,

            kegiatan_progress=kegiatan_progress,

            kegiatan_list=kegiatan_list,

            tahun=datetime.now().year,

            tahun_filter=tahun_filter,

            tahun_list=tahun_list
        )

    # PPL
    elif current_user.role == "ppl":

        bs_list = (
            BlokSensus.query
            .join(Kegiatan)
            .filter(
                BlokSensus.ppl_id == current_user.id,
                Kegiatan.status != "Selesai"
            )
            .all()
        )

        return render_template(
            "dashboard_ppl.html",
            bs_list=bs_list
        )

    # PML
    elif current_user.role == "pml":

        bs_list = (
            BlokSensus.query
            .join(Kegiatan)
            .filter(
                Kegiatan.status != "Selesai"
            )
            .all()
        )

        return render_template(
            "dashboard_pml.html",
            bs_list=bs_list
        )


# ==========================
# USERS
# ==========================

@app.route('/users')
@login_required
def users():

    users = User.query.all()

    print(users)
    print(len(users))

    return render_template(
        'users_list.html',
        users=users
    )

# ==========================
# TAMBAH USER
# ==========================

@app.route(
    "/users/tambah",
    methods=["GET", "POST"]
)
@login_required
def tambah_user():

    if current_user.role != "admin":
        return redirect(
            url_for("dashboard")
        )

    if request.method == "POST":

        user = User(

            nama=request.form["nama"],

            username=request.form["username"],

            password=generate_password_hash(

                request.form["password"]

            ),

            role=request.form["role"],

            no_hp=request.form["no_hp"],

            desa=request.form["desa"],

            kecamatan=request.form["kecamatan"]

        )

        db.session.add(user)
        db.session.commit()

        flash(
            "User berhasil dibuat",
            "success"
        )

        return redirect(
            url_for("users")
        )

    return render_template(
        "user_form.html",
        user=None
    )


# ==========================
# EDIT USER
# ==========================

@app.route(
    "/users/edit/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_user(id):

    if current_user.role != "admin":
        return redirect(
            url_for("dashboard")
        )

    user = User.query.get_or_404(id)

    if request.method == "POST":

        user.nama = request.form["nama"]
        user.username = request.form["username"]
        user.role = request.form["role"]
        user.no_hp = request.form["no_hp"]

        user.desa = request.form["desa"]

        user.kecamatan = request.form["kecamatan"]

        if request.form["password"]:

            user.password = (
                generate_password_hash(
                    request.form["password"]
                )
            )

        db.session.commit()

        flash(
            "User berhasil diubah",
            "success"
        )

        return redirect(
            url_for("users")
        )

    return render_template(
        "user_form.html",
        user=user
    )


# ==========================
# HAPUS USER
# ==========================

@app.route(
    "/users/hapus/<int:id>"
)
@login_required
def hapus_user(id):

    if current_user.role != "admin":
        return redirect(
            url_for("dashboard")
        )

    user = User.query.get_or_404(id)

    if user.username == "admin":

        flash(
            "Admin tidak boleh dihapus",
            "danger"
        )

        return redirect(
            url_for("users")
        )

    db.session.delete(user)
    db.session.commit()

    flash(
        "User berhasil dihapus",
        "success"
    )

    return redirect(
        url_for("users")
    )
    
# ==========================
# KEGIATAN
# ==========================

@app.route("/kegiatan")
@login_required
def kegiatan():

    data = Kegiatan.query.order_by(
        Kegiatan.tahun.desc()
    ).all()
    

    return render_template(
        "kegiatan_list.html",
        data=data
    )


@app.route(
    "/kegiatan/tambah",
    methods=["GET", "POST"]
)
@login_required
def tambah_kegiatan():

    if current_user.role != "admin":
        return redirect(
            url_for("dashboard")
        )

    if request.method == "POST":

        data = Kegiatan(
            nama=request.form["nama"],
            tahun=int(request.form["tahun"]),
            tanggal_mulai=datetime.strptime(
                request.form["tanggal_mulai"],
                "%Y-%m-%d"
            ).date(),
            tanggal_selesai=datetime.strptime(
                request.form["tanggal_selesai"],
                "%Y-%m-%d"
            ).date(),
            status=request.form["status"]
        )

        db.session.add(data)
        db.session.commit()

        flash(
            "Kegiatan berhasil ditambahkan",
            "success"
        )

        return redirect(
            url_for("kegiatan")
        )

    return render_template(
        "kegiatan_form.html",
        kegiatan=None
    )


@app.route(
    "/kegiatan/edit/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_kegiatan(id):

    data = Kegiatan.query.get_or_404(id)

    if request.method == "POST":

        data.nama = request.form["nama"]
        data.tahun = int(
            request.form["tahun"]
        )
        data.tanggal_mulai = datetime.strptime(
        request.form["tanggal_mulai"],
        "%Y-%m-%d"
        ).date()

        data.tanggal_selesai = datetime.strptime(
            request.form["tanggal_selesai"],
            "%Y-%m-%d"
        ).date()
        data.status = request.form["status"]

        db.session.commit()

        flash(
            "Kegiatan berhasil diubah",
            "success"
        )

        return redirect(
            url_for("kegiatan")
        )

    return render_template(
        "kegiatan_form.html",
        kegiatan=data
    )


@app.route("/kegiatan/hapus/<int:id>")
@login_required
def hapus_kegiatan(id):

    kegiatan = Kegiatan.query.get_or_404(id)

    BlokSensus.query.filter_by(
        kegiatan_id=id
    ).delete()

    Penugasan.query.filter_by(
        kegiatan_id=id
    ).delete()

    db.session.delete(kegiatan)
    db.session.commit()

    flash(
        "Kegiatan berhasil dihapus",
        "success"
    )

    return redirect(
        url_for("kegiatan")
    )
# ==========================
# PENUGASAN
# ==========================

@app.route("/penugasan")
@login_required
def penugasan():

    data = Penugasan.query.all()

    return render_template(
        "penugasan_list.html",
        data=data
    )


@app.route(
    "/penugasan/tambah",
    methods=["GET", "POST"]
)
@login_required
def tambah_penugasan():

    kegiatan = Kegiatan.query.all()

    # ambil semua user dulu
    semua_user = User.query.all()

    print("\n===== USER =====")
    for u in semua_user:
        print(u.id, u.nama, repr(u.role))

    ppl_list = User.query.filter_by(
        role="ppl"
    ).all()

    pml_list = User.query.filter_by(
        role="pml"
    ).all()

    print("Jumlah PPL :", len(ppl_list))
    print("Jumlah PML :", len(pml_list))


    if request.method == "POST":

        kegiatan_id = int(
            request.form["kegiatan_id"]
        )

        pml_id = int(
            request.form["pml_id"]
        )

        ppl_ids = request.form.getlist(
            "ppl_ids[]"
        )

        for ppl_id in ppl_ids:

            if ppl_id:

                data = Penugasan(

                    kegiatan_id=kegiatan_id,

                    pml_id=pml_id,

                    ppl_id=int(ppl_id)

                )

                db.session.add(data)

        db.session.commit()

        flash(
            "Penugasan berhasil dibuat",
            "success"
        )

        return redirect(
            url_for("penugasan")
        )


    return render_template(

        "penugasan_form.html",

        kegiatan=kegiatan,

        ppl_list=ppl_list,

        pml_list=pml_list

    )


@app.route(
    "/penugasan/hapus/<int:id>"
)
@login_required
def hapus_penugasan(id):

    data = Penugasan.query.get_or_404(id)

    db.session.delete(data)
    db.session.commit()

    flash(
        "Penugasan berhasil dihapus",
        "success"
    )

    return redirect(
        url_for("penugasan")
    )

# ==========================
# BLOK SENSUS
# ==========================

@app.route("/bs")
@login_required
def bs():

    data = BlokSensus.query.order_by(
        BlokSensus.kode_bs
    ).all()

    return render_template(
        "bs_list.html",
        data=data
    )


@app.route(
    "/bs/tambah",
    methods=["GET", "POST"]
)
@login_required
def tambah_bs():

    kegiatan = Kegiatan.query.all()

    ppl_list = User.query.filter_by(
        role="ppl"
    ).all()

    if request.method == "POST":

        daftar_bs = request.form["kode_bs"]

        for baris in daftar_bs.splitlines():

            baris = baris.strip()

            if not baris:
                continue

            try:

                kode_bs, target = baris.split(":")

                kode_bs = kode_bs.strip()
                target = int(target.strip())

            except ValueError:

                flash(
                    f"Format salah: {baris}",
                    "danger"
                )

                return redirect(
                    url_for("tambah_bs")
                )

            bs = BlokSensus(
                kegiatan_id=int(
                    request.form["kegiatan_id"]
                ),
                ppl_id=int(
                    request.form["ppl_id"]
                ),
                kecamatan=request.form["kecamatan"],
                kode_bs=kode_bs,
                target=target,
                realisasi=0,
                diperiksa=0,
                status="Belum"
            )

            db.session.add(bs)

        db.session.commit()

    return render_template(
        "bs_form.html",
        kegiatan=kegiatan,
        ppl_list=ppl_list,
        data=None
    )


@app.route(
    "/bs/edit/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_bs(id):

    data = BlokSensus.query.get_or_404(id)

    kegiatan = Kegiatan.query.all()

    ppl_list = User.query.filter_by(
        role="ppl"
    ).all()

    if request.method == "POST":

        data.kegiatan_id = int(
            request.form["kegiatan_id"]
        )

        data.ppl_id = int(
            request.form["ppl_id"]
        )
        data.kecamatan = request.form["kecamatan"]

        data.kode_bs = request.form["kode_bs"]

        data.target = int(
            request.form["target"]
        )

        db.session.commit()

        flash(
            "BS berhasil diubah",
            "success"
        )

        return redirect(
            url_for("bs")
        )

    return render_template(
        "bs_form.html",
        kegiatan=kegiatan,
        ppl_list=ppl_list,
        data=data
    )


@app.route(
    "/bs/hapus/<int:id>"
)
@login_required
def hapus_bs(id):

    data = BlokSensus.query.get_or_404(id)

    db.session.delete(data)
    db.session.commit()

    flash(
        "BS berhasil dihapus",
        "success"
    )

    return redirect(
        url_for("bs")
    )
@app.route(
    "/bs/update/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def update_bs(id):

    data = BlokSensus.query.get_or_404(id)

    if request.method == "POST":

        realisasi = int(
            request.form["realisasi"]
        )

        data.realisasi = realisasi

        if realisasi == 0:
            data.status = "Belum"
        elif realisasi < data.target:
            data.status = "Berjalan"
        else:
            data.status = "Selesai"

        db.session.commit()

        return redirect(
            url_for("dashboard")
        )

    return render_template(
        "update_bs.html",
        data=data
    )

@app.route(
    "/pml/periksa/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def periksa_bs(id):

    data = BlokSensus.query.get_or_404(id)

    if request.method == "POST":

        data.diperiksa = int(
            request.form["diperiksa"]
        )


        db.session.commit()

        flash(
            "Pemeriksaan berhasil disimpan",
            "success"
        )

        return redirect(
            url_for("dashboard")
        )

    return render_template(
        "periksa_bs.html",
        data=data
    )
@app.route("/kegiatan/detail/<int:id>")
@login_required
def detail_kegiatan(id):

    kegiatan = Kegiatan.query.get_or_404(id)

    bs_list = BlokSensus.query.filter_by(
        kegiatan_id=id
    ).all()

    penugasan = Penugasan.query.filter_by(
        kegiatan_id=id
    ).all()

    mapping_pml = {
        p.ppl_id: p.pml.nama
        for p in penugasan
    }

    progress_kecamatan = {}

    for bs in bs_list:

        kec = bs.kecamatan

        if kec not in progress_kecamatan:
            progress_kecamatan[kec] = {
                "target": 0,
                "realisasi": 0,
                "diperiksa": 0
            }

        progress_kecamatan[kec]["target"] += bs.target
        progress_kecamatan[kec]["realisasi"] += bs.realisasi
        progress_kecamatan[kec]["diperiksa"] += bs.diperiksa

    for item in progress_kecamatan.values():

        if item["target"] > 0:
            item["progress_ppl"] = round(
                item["realisasi"] * 100 / item["target"], 1
            )

            item["progress_pml"] = round(
                item["diperiksa"] * 100 / item["target"], 1
            )
        else:
            item["progress_ppl"] = 0
            item["progress_pml"] = 0

    return render_template(
        "kegiatan_detail.html",
        kegiatan=kegiatan,
        bs_list=bs_list,
        mapping_pml=mapping_pml,
        progress_kecamatan=progress_kecamatan
    )
        
@app.route(
    "/kegiatan/<int:kegiatan_id>/anomali/tambah",
    methods=["GET", "POST"]
)
def tambah_anomali(kegiatan_id):

    kegiatan = Kegiatan.query.get_or_404(
        kegiatan_id
    )

    if request.method == "POST":

        anomali = LinkAnomali(
            kegiatan_id=kegiatan.id,
            judul=request.form["judul"],
            link=request.form["link"],
            keterangan=request.form["keterangan"]
        )

        db.session.add(anomali)
        db.session.commit()

        return redirect(
            url_for(
                "detail_kegiatan",
                id=kegiatan.id
            )
        )

    return render_template(
        "anomali_form.html",
        kegiatan=kegiatan
    )

@app.route("/anomali/hapus/<int:id>")
def hapus_anomali(id):

    anomali = LinkAnomali.query.get_or_404(id)

    kegiatan_id = anomali.kegiatan_id

    db.session.delete(anomali)
    db.session.commit()

    return redirect(
        url_for(
            "detail_kegiatan",
            id=kegiatan_id
        )
    )

@app.route(
    "/kegiatan/<int:id>/drive",
    methods=["GET", "POST"]
)
def edit_drive(id):

    kegiatan = Kegiatan.query.get_or_404(id)

    if request.method == "POST":

        kegiatan.drive_link = request.form["drive_link"]

        db.session.commit()

        return redirect(
            url_for(
                "detail_kegiatan",
                id=id
            )
        )

    return render_template(
        "drive_form.html",
        kegiatan=kegiatan
    )
    
    
from ai.predictor import (
    cari_kbli,
    cari_kbji
)


# ==========================
# KBLI AI
# ==========================

@app.route("/kbli")
@login_required
def kbli():

    return render_template(
        "kbli.html"
    )


@app.route(
    "/kbli/prediksi",
    methods=["GET", "POST"]
)
@login_required
def prediksi_kbli():

    if request.method == "GET":

        return redirect(
            url_for("kbli")
        )

    uraian = request.form.get(
        "uraian",
        ""
    ).strip()

    if not uraian:

        flash(
            "Uraian tidak boleh kosong",
            "danger"
        )

        return redirect(
            url_for("kbli")
        )

    hasil_kbli = cari_kbli(
        uraian
    )

    hasil_kbji = cari_kbji(
        uraian
    )

    return render_template(
        "hasil_kbli.html",
        uraian=uraian,
        hasil_kbli=hasil_kbli,
        hasil_kbji=hasil_kbji
    )

@app.route(
    "/kegiatan/<int:kegiatan_id>/link-lainnya/tambah",
    methods=["GET", "POST"]
)
@login_required
def tambah_link_lainnya(kegiatan_id):

    kegiatan = Kegiatan.query.get_or_404(
        kegiatan_id
    )

    if request.method == "POST":

        data = LinkLainnya(
            kegiatan_id=kegiatan.id,
            judul=request.form["judul"],
            link=request.form["link"],
            keterangan=request.form["keterangan"]
        )

        db.session.add(data)
        db.session.commit()

        return redirect(
            url_for(
                "detail_kegiatan",
                id=kegiatan.id
            )
        )

    return render_template(
        "anomali_form.html",
        kegiatan=kegiatan
    )
@app.route(
    "/link-lainnya/hapus/<int:id>"
)
@login_required
def hapus_link_lainnya(id):

    data = LinkLainnya.query.get_or_404(id)

    kegiatan_id = data.kegiatan_id

    db.session.delete(data)
    db.session.commit()

    return redirect(
        url_for(
            "detail_kegiatan",
            id=kegiatan_id
        )
    )
    
@app.route("/forms")
@login_required
def forms():

    data = FormTemplate.query.all()

    return render_template(

        "form_list.html",

        data=data

    )
@app.route(
    "/forms/tambah",
    methods=["GET", "POST"]
)
@login_required
def tambah_form():

    if request.method == "POST":

        deadline = None

        if request.form.get("deadline"):
            deadline = datetime.strptime(
                request.form["deadline"],
                "%Y-%m-%d"
            ).date()


        data = FormTemplate(

            nama=request.form["nama"],

            role=request.form["role"],

            deadline=deadline

        )


        db.session.add(data)
        db.session.commit()

        flash(
            "Form berhasil ditambahkan",
            "success"
        )

        return redirect(
            url_for("forms")
        )


    return render_template(
        "form_form.html",
        form=None
    )
@app.route(

"/forms/<int:id>/field"

)

@login_required
def field(id):


    data = FormField.query.filter_by(

        form_id=id

    ).all()


    form = FormTemplate.query.get_or_404(

        id

    )


    return render_template(

        "field_list.html",

        data=data,

        form=form

    )
@app.route(

"/forms/<int:id>/field/tambah",

methods=["GET","POST"]

)

@login_required
def tambah_field(id):


    if request.method=="POST":



        data = FormField(


            form_id=id,


            pertanyaan=request.form[

                "pertanyaan"

            ],


            tipe=request.form[

                "tipe"

            ]

        )


        db.session.add(

            data

        )


        db.session.commit()



        return redirect(

            url_for(

                "field",

                id=id

            )

        )



    return render_template(

        "field_form.html"

    )
@app.route('/my-form/<int:id>')
@login_required
def lihat_isian(id):

    form = FormTemplate.query.get_or_404(id)

    fields = FormField.query.filter_by(
        form_id=id
    ).all()


    responses = FormResponse.query.filter_by(

        form_id=id,

        user_id=current_user.id

    ).all()


    rows = []

    for r in responses:

        row = {

            'id': r.id

        }


        answers = FormAnswer.query.filter_by(

            response_id=r.id

        ).all()


        for a in answers:

            row[a.field_id] = a.jawaban


        rows.append(

            row

        )


    return render_template(

        'lihat_isian.html',

        form=form,

        fields=fields,

        rows=rows

    )

@app.route(

"/isi-form/<int:id>",

methods=["GET","POST"]

)

@login_required
def isi_form(id):



    form = FormTemplate.query.get_or_404(

        id

    )



    fields = FormField.query.filter_by(

        form_id=id

    ).all()



    if request.method=="POST":

        response = FormResponse(

            form_id=id,

            user_id=current_user.id

        )

        db.session.add(response)
        db.session.commit()


        for f in fields:

            ans = FormAnswer(

                response_id=response.id,

                field_id=f.id,

                jawaban=request.form.get(

                    str(f.id)

                )

            )

            db.session.add(ans)


        db.session.commit()


        flash(

            "Data berhasil ditambahkan",

            "success"

        )


        return redirect(

            url_for(

                "isi_form",

                id=id

            )

        )



    return render_template(

        "isi_form.html",

        form=form,

        fields=fields

    )
        
@app.route("/forms/<int:id>/hasil")
@login_required
def hasil_form(id):

    form = FormTemplate.query.get_or_404(id)

    users = User.query.filter_by(
        role=form.role
    ).order_by(
        User.nama
    ).all()


    responses = FormResponse.query.filter_by(
        form_id=id
    ).all()


    sudah = {
        r.user_id:r
        for r in responses
    }


    return render_template(

        "hasil_form.html",

        form=form,

        users=users,

        sudah=sudah

    )



@app.route(
"/forms/<int:id>/detail/<int:user_id>"
)

@login_required
def detail_respon(

id,

user_id

):


    form = FormTemplate.query.get_or_404(

        id

    )


    user = User.query.get_or_404(

        user_id

    )


    response = FormResponse.query.filter_by(

        form_id=id,

        user_id=user_id

    ).first()



    if not response:

        flash(

            "Belum ada jawaban",

            "warning"

        )

        return redirect(

            url_for(

                "hasil_form",

                id=id

            )

        )



    answers = FormAnswer.query.filter_by(

        response_id=response.id

    ).all()



    mapping = {}


    for a in answers:


        field = FormField.query.get(

            a.field_id

        )


        mapping[

            field.pertanyaan

        ] = a.jawaban



    return render_template(

        "detail_respon.html",

        form=form,

        user=user,

        mapping=mapping

    )
@app.route(
"/forms/field/edit/<int:id>",
methods=["GET","POST"]
)
@login_required
def edit_field(id):

    field = FormField.query.get_or_404(id)

    if request.method=="POST":

        field.pertanyaan = request.form[
            "pertanyaan"
        ]

        field.tipe = request.form[
            "tipe"
        ]

        db.session.commit()

        flash(

            "Pertanyaan berhasil diubah",

            "success"

        )

        return redirect(

            url_for(

                "field",

                id=field.form_id

            )

        )


    return render_template(

        "field_form.html",

        field=field

    )
@app.route(
"/forms/field/hapus/<int:id>"
)
@login_required
def hapus_field(id):


    field = FormField.query.get_or_404(id)

    form_id = field.form_id


    db.session.delete(

        field

    )


    db.session.commit()


    flash(

        "Pertanyaan dihapus",

        "success"

    )


    return redirect(

        url_for(

            "field",

            id=form_id

        )

    )

@app.route("/forms/<int:id>/excel")
@login_required
def export_excel(id):


    form = FormTemplate.query.get_or_404(id)


    fields = FormField.query.filter_by(

        form_id=id

    ).order_by(

        FormField.id

    ).all()



    wb = Workbook()

    ws = wb.active



    ws.title = form.nama



    header = ["Nama"]



    for f in fields:

        header.append(

            f.pertanyaan

        )



    ws.append(

        header

    )



    responses = FormResponse.query.filter_by(

        form_id=id

    ).all()




    for r in responses:



        user = User.query.get(

            r.user_id

        )



        row = [

            user.nama

            if user

            else

            "User Dihapus"

        ]




        for f in fields:



            ans = FormAnswer.query.filter_by(


                response_id=r.id,


                field_id=f.id


            ).first()




            if ans:


                row.append(

                    ans.jawaban

                )



            else:


                row.append("")



        ws.append(

            row

        )



    output = BytesIO()


    wb.save(

        output

    )


    output.seek(

        0

    )



    return send_file(


        output,


        download_name=f"{form.nama}.xlsx",


        as_attachment=True,


        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


    )

@app.route(
    "/forms/hapus/<int:id>"
)
@login_required
def hapus_form(id):

    form = FormTemplate.query.get_or_404(id)

    fields = FormField.query.filter_by(
        form_id=id
    ).all()

    responses = FormResponse.query.filter_by(
        form_id=id
    ).all()

    for r in responses:

        FormAnswer.query.filter_by(
            response_id=r.id
        ).delete()

    FormResponse.query.filter_by(
        form_id=id
    ).delete()

    FormField.query.filter_by(
        form_id=id
    ).delete()

    db.session.delete(form)

    db.session.commit()

    flash(
        "Form berhasil dihapus",
        "success"
    )

    return redirect(
        url_for("forms")
    )
@app.route("/dashboard-strategis")
@login_required
def dashboard_strategis():

    kategori = DashboardKategori.query.order_by(
        DashboardKategori.nama
    ).all()

    return render_template(
        "dashboard_strategis.html",
        kategori=kategori
    )
@app.route("/dashboard-strategis/<int:id>")
@login_required
def indikator_strategis(id):

    kategori = DashboardKategori.query.get_or_404(id)

    indikator = DashboardIndikator.query.filter_by(
        kategori_id=id
    ).all()

    return render_template(
        "indikator_strategis.html",
        kategori=kategori,
        indikator=indikator
    )
@app.route(
    "/dashboard-strategis/kategori/tambah",
    methods=["GET","POST"]
)
@login_required
def tambah_kategori():

    if request.method == "POST":

        data = DashboardKategori(
            nama=request.form["nama"]
        )

        db.session.add(data)
        db.session.commit()

        return redirect(
            url_for(
                "dashboard_strategis"
            )
        )

    return render_template(
        "kategori_form.html"
    )
@app.route(
    "/dashboard-strategis/indikator/tambah",
    methods=["GET","POST"]
)
@login_required
def tambah_indikator():

    kategori = DashboardKategori.query.all()

    if request.method == "POST":

        data = DashboardIndikator(

            kategori_id=int(
                request.form["kategori_id"]
            ),

            nama=request.form["nama"],

            satuan=request.form["satuan"]

        )

        db.session.add(data)
        db.session.commit()

        return redirect(
            url_for(
                "dashboard_strategis"
            )
        )

    return render_template(
        "indikator_form.html",
        kategori=kategori
    )
@app.route(
    "/dashboard-strategis/data/<int:id>",
    methods=["GET","POST"]
)
@login_required
def input_data_strategis(id):

    indikator = DashboardIndikator.query.get_or_404(id)

    if request.method == "POST":

        data = DashboardData(

            indikator_id=id,

            tahun=int(
                request.form["tahun"]
            ),

            nilai=float(
                request.form["nilai"]
            )
        )

        db.session.add(data)
        db.session.commit()

        return redirect(
            url_for(
                "indikator_strategis",
                id=indikator.kategori_id
            )
        )

    return render_template(
        "data_form.html",
        indikator=indikator
    )

@app.route(
"/dashboard-strategis/grafik/<int:id>"
)
@login_required
def grafik_indikator(id):


    indikator = DashboardIndikator.query.get_or_404(

        id

    )


    tahun_list = db.session.query(

        DashboardData.tahun

    ).filter_by(

        indikator_id=id

    ).distinct().all()



    tahun_list = sorted(

        [x[0] for x in tahun_list]

    )



    tahun = request.args.get(

        "tahun",

        type=int

    )



    if tahun_list:


        if not tahun:


            tahun = tahun_list[-1]


    else:


        tahun = None



    data=[]

    labels=[]

    values=[]



    if tahun:


        data = DashboardData.query.filter_by(

            indikator_id=id,

            tahun=tahun

        ).all()



        labels = [

            x.wilayah

            for x in data

        ]



        values = [

            x.nilai

            for x in data

        ]



    return render_template(

        "chart_view.html",


        indikator=indikator,


        tahun=tahun,


        tahun_list=tahun_list,


        labels=labels,


        values=values

    )
@app.route(

"/dashboard/upload/<int:id>",

methods=[

"POST"

]

)

def upload_dashboard(id):


    file=request.files['file']


    df=pd.read_excel(

        file

    )


    for _,r in df.iterrows():



        x=DashboardData(

            indikator_id=id,


            tahun=int(

                r['tahun']

            ),


            nilai=float(

                r['nilai']

            )

        )


        db.session.add(

            x

        )



    db.session.commit()



    return redirect(

        url_for(

            'grafik_indikator',

            id=id

        )

    )
@app.route(
'/dashboard-sheet/<int:id>',
methods=['GET','POST']
)
@login_required
def dashboard_sheet(id):

    indikator = DashboardIndikator.query.get_or_404(id)

    baris = DashboardBaris.query.filter_by(
        indikator_id=id
    ).all()

    kolom = DashboardKolom.query.filter_by(
        indikator_id=id
    ).all()


    if request.method == 'POST':

        for b in baris:

            b.nama = request.form.get(
                f'baris_{b.id}'
            )


        for k in kolom:

            k.nama = request.form.get(
                f'kolom_{k.id}'
            )


        for b in baris:

            for k in kolom:

                nilai = request.form.get(

                    f'{b.id}_{k.id}'

                )


                ada = DashboardData.query.filter_by(

                    indikator_id=id,

                    wilayah=b.nama,

                    tahun=int(k.nama)

                ).first()


                if nilai:


                    if ada:

                        ada.nilai=float(nilai)

                    else:


                        db.session.add(

                            DashboardData(

                                indikator_id=id,

                                wilayah=b.nama,

                                tahun=int(k.nama),

                                nilai=float(nilai)

                            )

                        )


        db.session.commit()



        return redirect(

        url_for(

        'grafik_indikator',

        id=id

        )

        )



    data={}


    semua=DashboardData.query.filter_by(

        indikator_id=id

    ).all()


    for d in semua:

        data[(d.wilayah,d.tahun)] = d.nilai



    return render_template(

    'sheet_dashboard.html',

    indikator=indikator,

    baris=baris,

    kolom=kolom,

    data=data

    )








@app.route(
'/dashboard/export/<int:id>'
)
@login_required
def export_dashboard(id):


    indikator = DashboardIndikator.query.get_or_404(id)


    data = DashboardData.query.filter_by(

        indikator_id=id

    ).all()


    wilayah = sorted(

        list(

            set(

                x.wilayah

                for x in data

            )

        )

    )


    tahun = sorted(

        list(

            set(

                x.tahun

                for x in data

            )

        )

    )


    wb = Workbook()

    ws = wb.active


    ws.title = indikator.nama



    judul = f"{indikator.nama} ({indikator.satuan})"


    ws["A1"] = judul


    ws["A1"].font = Font(

        bold=True,

        size=14

    )


    ws["A1"].alignment = Alignment(

        horizontal="center"

    )



    ws.merge_cells(

        start_row=1,

        start_column=1,

        end_row=1,

        end_column=len(tahun)+1

    )



    header = [

        "Kecamatan"

    ] + tahun



    for i,v in enumerate(

        header,

        1

    ):


        ws.cell(

            row=3,

            column=i

        ).value = v



        ws.cell(

            row=3,

            column=i

        ).font = Font(

            bold=True

        )



    row_excel = 4


    for k in wilayah:


        row=[

            k

        ]


        for t in tahun:



            d = DashboardData.query.filter_by(


                indikator_id=id,


                wilayah=k,


                tahun=t


            ).first()



            row.append(

                d.nilai

                if d

                else ''

            )



        for i,v in enumerate(

            row,

            1

        ):


            ws.cell(

                row=row_excel,

                column=i

            ).value = v



        row_excel += 1



    output = BytesIO()


    wb.save(

        output

    )


    output.seek(

        0

    )


    return send_file(

        output,


        download_name=f'{indikator.nama}.xlsx',


        as_attachment=True,


        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )
@app.route('/dashboard-sheet/<int:id>/baris/<int:jumlah>')
@login_required
def tambah_baris(id, jumlah):

    for i in range(jumlah):

        db.session.add(

            DashboardBaris(

                indikator_id=id,

                nama=f'Baris {i+1}'

            )

        )

    db.session.commit()

    return redirect(

        url_for(

            'dashboard_sheet',

            id=id

        )

    )


@app.route('/dashboard-sheet/<int:id>/kolom/<int:jumlah>')
@login_required
def tambah_kolom(id, jumlah):

    awal = DashboardKolom.query.filter_by(

        indikator_id=id

    ).count()


    for i in range(jumlah):

        db.session.add(

            DashboardKolom(

                indikator_id=id,

                nama=str(

                    2026 +

                    awal +

                    i

                )

            )

        )



    db.session.commit()


    return redirect(

        url_for(

            'dashboard_sheet',

            id=id

        )

    )


@app.route('/hapus-baris/<int:id>')
@login_required
def hapus_baris(id):

    x = DashboardBaris.query.get_or_404(id)

    db.session.delete(x)

    db.session.commit()

    return redirect(request.referrer)


@app.route('/hapus-kolom/<int:id>')
@login_required
def hapus_kolom(id):

    x = DashboardKolom.query.get_or_404(id)

    db.session.delete(x)

    db.session.commit()

    return redirect(request.referrer)

@app.route(
'/dashboard/indikator/hapus/<int:id>'
)

@login_required
def hapus_indikator(id):


    indikator = DashboardIndikator.query.get_or_404(id)



    DashboardData.query.filter_by(

        indikator_id=id

    ).delete()



    DashboardBaris.query.filter_by(

        indikator_id=id

    ).delete()



    DashboardKolom.query.filter_by(

        indikator_id=id

    ).delete()



    kategori_id = indikator.kategori_id



    db.session.delete(

        indikator

    )



    db.session.commit()



    return redirect(

        url_for(

            'indikator_strategis',

            id=kategori_id

        )

    )

@login_required
def riwayat_form(id):


    responses = FormResponse.query.filter_by(

        form_id=id,

        user_id=current_user.id

    ).all()



    return render_template(

        'riwayat_form.html',

        responses=responses,

        form_id=id

    )
@login_required
def detail_isian(id):


    response = FormResponse.query.get_or_404(id)



    answers = FormAnswer.query.filter_by(

        response_id=id

    ).all()



    return render_template(

        'detail_isian.html',

        answers=answers

    )
@login_manager.user_loader
def load_user(user_id):

    try:

        return db.session.get(

            User,

            int(user_id)

        )

    except Exception:

        db.session.rollback()

        db.session.remove()

        return None

@app.route("/my-forms")
@login_required
def my_forms():

    data = FormTemplate.query.filter_by(

        role=current_user.role

    ).all()

    return render_template(

        "my_forms.html",

        data=data

    )
if __name__ == "__main__":
    app.run(debug=True)
